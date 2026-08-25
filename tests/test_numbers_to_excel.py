"""Apple Numbers -> Excel conversion.

Renaming a `.numbers` file to `.xlsx` gets you a workbook Excel calls corrupt, because
the bundle holds Apple's IWA archives rather than Excel's XML. These tests build real
Numbers documents with numbers-parser and pin what survives the trip into xlsx: cell
values, merged ranges, header rows, and Excel's own naming rules for worksheets.
"""

import datetime
import shutil
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "datagolf"))

numbers_parser = pytest.importorskip("numbers_parser")

from numbers_to_excel import (  # noqa: E402
    MAX_SHEET_TITLE,
    classify,
    convert,
    main,
    reject_reason,
    sheet_title,
)

ROUNDS = [
    ("Ludvig Åberg", 1.2345, datetime.datetime(2025, 7, 20), True),
    ("Scottie Scheffler", 2.5, datetime.datetime(2025, 7, 20), True),
    ("Nicolai Højgaard", -0.75, datetime.datetime(2025, 7, 20), False),
]


@pytest.fixture
def pga_numbers(tmp_path):
    """A Numbers document shaped like a season's raw PGA export."""
    doc = numbers_parser.Document(
        sheet_name="Raw 2025", table_name="Rounds", num_rows=len(ROUNDS) + 1, num_cols=4
    )
    table = doc.sheets[0].tables[0]
    for col, header in enumerate(["Player", "SG Total", "Round Date", "Made Cut"]):
        table.write(0, col, header)
    for row, values in enumerate(ROUNDS, start=1):
        for col, value in enumerate(values):
            table.write(row, col, value)

    doc.add_sheet("Notes", "Source", num_rows=2, num_cols=2)
    notes = doc.sheets[1].tables[0]
    notes.write(0, 0, "source")
    notes.write(0, 1, "=SUM(A1:A2) as text")
    notes.merge_cells("A2:B2")
    notes.write(1, 0, "merged note")

    path = tmp_path / "raw_data_pga_2025.numbers"
    doc.save(path)
    return path


def test_renaming_alone_does_not_produce_a_readable_workbook(pga_numbers):
    renamed = pga_numbers.with_suffix(".xlsx")
    shutil.copy(pga_numbers, renamed)
    with pytest.raises(Exception):
        openpyxl.load_workbook(renamed)


def test_converted_workbook_keeps_every_value(pga_numbers, tmp_path):
    convert(pga_numbers, tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(tmp_path / "out.xlsx")

    assert wb.sheetnames == ["Raw 2025", "Notes"]
    ws = wb["Raw 2025"]
    assert [c.value for c in ws[1]] == ["Player", "SG Total", "Round Date", "Made Cut"]
    assert ws.max_row == len(ROUNDS) + 1
    for row, expected in enumerate(ROUNDS, start=2):
        assert tuple(ws.cell(row=row, column=c).value for c in range(1, 5)) == expected


def test_header_row_is_bold_and_frozen(pga_numbers, tmp_path):
    convert(pga_numbers, tmp_path / "out.xlsx")
    ws = openpyxl.load_workbook(tmp_path / "out.xlsx")["Raw 2025"]

    assert ws["A1"].font.bold
    assert not ws["A2"].font.bold
    assert ws.freeze_panes == "B2"  # one header row, one header column


def test_merged_range_and_leading_equals_text_survive(pga_numbers, tmp_path):
    convert(pga_numbers, tmp_path / "out.xlsx")
    ws = openpyxl.load_workbook(tmp_path / "out.xlsx")["Notes"]

    assert "A2:B2" in [str(r) for r in ws.merged_cells.ranges]
    assert ws["A2"].value == "merged note"
    # Excel would evaluate this as a formula if the cell were not marked as text.
    assert ws["B1"].value == "=SUM(A1:A2) as text"
    assert ws["B1"].data_type == "s"


def test_default_output_sits_next_to_the_input(pga_numbers, capsys):
    assert main([str(pga_numbers)]) == 0
    capsys.readouterr()
    assert (pga_numbers.parent / "raw_data_pga_2025.xlsx").is_file()


def test_existing_output_needs_overwrite(pga_numbers, tmp_path, capsys):
    out = tmp_path / "out.xlsx"
    out.write_bytes(b"keep me")
    assert main([str(pga_numbers), "-o", str(out)]) == 1
    assert "--overwrite" in capsys.readouterr().err
    assert out.read_bytes() == b"keep me"

    assert main([str(pga_numbers), "-o", str(out), "--overwrite"]) == 0
    assert openpyxl.load_workbook(out).sheetnames == ["Raw 2025", "Notes"]


def test_a_numbers_bundle_already_renamed_to_xlsx_still_converts(pga_numbers, tmp_path):
    """The usual first attempt is renaming to .xlsx, so accept the file in that state."""
    misnamed = tmp_path / "raw_data_pga_2025_renamed.xlsx"
    shutil.copy(pga_numbers, misnamed)
    assert classify(misnamed) == "numbers"
    assert reject_reason(classify(misnamed), misnamed) is None

    out = tmp_path / "from_misnamed.xlsx"
    convert(misnamed, out)
    ws = openpyxl.load_workbook(out)["Raw 2025"]
    assert ws["A2"].value == ROUNDS[0][0]


def test_an_xlsx_renamed_to_numbers_is_reported_not_mangled(tmp_path):
    xlsx = tmp_path / "workbook.numbers"
    openpyxl.Workbook().save(xlsx)

    assert classify(xlsx) == "xlsx"
    reason = reject_reason("xlsx", xlsx)
    assert "already an Excel workbook" in reason and "workbook.xlsx" in reason


@pytest.mark.parametrize("kind", ["xls", "ods", "numbers-legacy", "unknown"])
def test_every_rejected_format_explains_itself(kind, tmp_path):
    assert reject_reason(kind, tmp_path / "f.numbers")


def test_missing_input_fails_without_a_traceback(tmp_path, capsys):
    assert main([str(tmp_path / "nope.numbers")]) == 1
    assert "not found" in capsys.readouterr().err


@pytest.mark.parametrize("name,expected", [
    ("Rounds", "Rounds"),
    ("Q1/Q2: raw [2025]?*", "Q1-Q2- raw -2025---"),
    ("  spaced   out  ", "spaced out"),
    ("", "Table"),
])
def test_worksheet_names_obey_excel_rules(name, expected):
    assert sheet_title(name, set()) == expected


def test_long_worksheet_names_are_clipped_and_stay_unique():
    used = set()
    long_name = "PGA 2025 raw shot level data export"
    first = sheet_title(long_name, used)
    second = sheet_title(long_name, used)

    assert first == long_name[:MAX_SHEET_TITLE]
    assert second != first and len(second) <= MAX_SHEET_TITLE
    assert sheet_title("rounds", {"rounds"}) == "rounds (2)"  # Excel names are case-insensitive
