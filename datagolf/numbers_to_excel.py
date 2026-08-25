#!/usr/bin/env python3
"""
Convert Apple Numbers spreadsheets (`.numbers`) to Excel workbooks (`.xlsx`).

A `.numbers` file is not a spreadsheet document, it is a zipped bundle of Apple's
IWA (protobuf) archives, so renaming it to `.xlsx` does not make Excel able to
read it — Excel reports the file as corrupt. The data has to be parsed out and
re-written, which is what this script does, with no Mac or Numbers install needed.

Every table in the Numbers document becomes one worksheet, keeping cell values
(text, numbers, dates, durations, booleans), merged ranges, header rows/columns
and column widths. Formulas are exported as the values Numbers last calculated,
because Numbers' formula dialect does not always map onto Excel's.

Inputs are identified by content, not extension, so a bundle already renamed to
`.xlsx` in the hope Excel would read it converts fine, and a real Excel workbook
misnamed `.numbers` is reported rather than mangled.

Usage:
    # writes raw_data_pga_2025.xlsx next to the input
    python datagolf/numbers_to_excel.py raw_data_pga_2025.numbers

    # explicit destination, or a directory for several inputs at once
    python datagolf/numbers_to_excel.py in.numbers -o datagolf/workbooks/out.xlsx
    python datagolf/numbers_to_excel.py *.numbers --outdir datagolf/workbooks

    # inspect the sheets/tables without writing anything
    python datagolf/numbers_to_excel.py raw_data_pga_2025.numbers --list

Requires `numbers-parser` and `openpyxl` (both in the repo's requirements.txt).
"""

import argparse
import contextlib
import itertools
import re
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Excel's own limits. Exceeding any of them produces a file Excel refuses to open.
MAX_SHEET_TITLE = 31
MAX_CELL_TEXT = 32767
MAX_ROWS = 1048576
MAX_COLS = 16384

# Characters Excel does not allow in a worksheet name.
INVALID_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")

# Numbers stores column widths in points; Excel counts them in default-font characters.
POINTS_PER_CHARACTER = 7.0


def classify(path: Path) -> str:
    """Identify a file by its contents rather than its extension.

    Numbers documents get renamed and mailed around a lot, so the extension is
    routinely wrong in both directions: an xlsx saved as `.numbers`, or a real
    Numbers bundle already renamed to `.xlsx` in the hope that Excel would open it.
    """
    if path.is_dir():
        names = {p.name for p in path.iterdir()}
        if names & {"index.xml", "index.xml.gz"}:
            return "numbers-legacy"
        if names & {"Index", "Index.zip", "Metadata"}:
            return "numbers"
        return "unknown"

    with open(path, "rb") as f:
        magic = f.read(8)
    if magic.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "xls"  # OLE2: Excel 97-2003 and other legacy Office documents
    if not magic.startswith(b"PK"):
        return "unknown"

    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
    except zipfile.BadZipFile:
        return "unknown"
    if any(n.startswith("xl/") for n in names):
        return "xlsx"
    if any(n == "Index.zip" or n.startswith("Index/") for n in names):
        return "numbers"
    if any(n.startswith("Metadata/") for n in names):
        return "numbers"
    if any(n in ("index.xml", "index.xml.gz") for n in names):
        return "numbers-legacy"
    if "content.xml" in names:
        return "ods"
    return "zip"


def reject_reason(kind: str, path: Path) -> str | None:
    """Why `path` cannot be converted, or None when it looks like a Numbers document."""
    if kind == "numbers":
        return None
    if kind == "xlsx":
        return (f"{path.name} is already an Excel workbook stored under a .numbers name. "
                f"Rename it to {path.with_suffix('.xlsx').name} and Excel will open it as is.")
    if kind == "numbers-legacy":
        return (f"{path.name} is an iWork '09 era Numbers document, which this converter "
                "cannot read. Open it in Numbers once and re-save, or export to Excel there.")
    if kind == "xls":
        return (f"{path.name} is a legacy Excel/Office (OLE2) file, not a Numbers document. "
                "Open it in Excel and use Save As to get an .xlsx.")
    if kind == "ods":
        return f"{path.name} is an OpenDocument spreadsheet, not a Numbers document."
    return (f"{path.name} is not an Apple Numbers document (unrecognised format). "
            "Check the file downloaded completely.")


def sheet_title(base: str, used: set) -> str:
    """An Excel-legal, unique worksheet name derived from `base`."""
    title = INVALID_TITLE_CHARS.sub("-", str(base or ""))
    title = re.sub(r"\s+", " ", title).strip().strip("'")
    title = (title or "Table")[:MAX_SHEET_TITLE]
    if title.casefold() not in used:
        used.add(title.casefold())
        return title
    for n in itertools.count(2):
        suffix = f" ({n})"
        candidate = title[:MAX_SHEET_TITLE - len(suffix)].rstrip() + suffix
        if candidate.casefold() not in used:
            used.add(candidate.casefold())
            return candidate


def table_titles(doc) -> list:
    """(sheet, table, worksheet_title) for every table, in document order.

    A sheet holding a single table lends it its own name, which is what people
    recognise; sheets with several tables get "Sheet - Table" names.
    """
    used = set()
    out = []
    for sheet in doc.sheets:
        tables = list(sheet.tables)
        for table in tables:
            base = sheet.name if len(tables) == 1 else f"{sheet.name} - {table.name}"
            out.append((sheet, table, sheet_title(base, used)))
    return out


@contextlib.contextmanager
def numbers_suffixed(path: Path):
    """Yield `path` under a `.numbers` suffix, which numbers-parser insists on.

    The usual first attempt at this conversion is to rename the bundle to `.xlsx`
    and hope Excel copes, so the file often arrives with the wrong extension while
    its contents are still a Numbers document. Stage a copy rather than making
    someone rename it back.
    """
    if path.suffix == ".numbers":
        yield path
        return
    with tempfile.TemporaryDirectory(prefix="numbers-to-excel-") as tmp:
        staged = Path(tmp) / (path.stem + ".numbers")
        if path.is_dir():
            shutil.copytree(path, staged)
        else:
            shutil.copy2(path, staged)
        yield staged


def cell_value(cell):
    """The Excel-safe value of a Numbers cell, or None for empty/unrepresentable cells.

    Returns (value, was_truncated) so the caller can report text clipped to Excel's
    32,767 character cell limit.
    """
    value = getattr(cell, "value", None)
    if not isinstance(value, str):
        return value, False
    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    if len(value) > MAX_CELL_TEXT:
        return value[:MAX_CELL_TEXT], True
    return value, False


def write_table(ws, table) -> tuple:
    """Copy one Numbers table into an empty worksheet. Returns (cells_written, truncated)."""
    written = truncated = 0
    for r, row in enumerate(table.rows(), start=1):
        for c, cell in enumerate(row, start=1):
            value, clipped = cell_value(cell)
            if value is None:
                continue
            out = ws.cell(row=r, column=c)
            out.value = value
            # openpyxl treats a leading "=" as a formula; Numbers text stays text.
            if isinstance(value, str) and out.data_type == "f":
                out.data_type = "s"
            written += 1
            truncated += clipped

    for cell_range in table.merge_ranges:
        ws.merge_cells(cell_range)

    header_rows = min(table.num_header_rows or 0, table.num_rows)
    header_cols = min(table.num_header_cols or 0, table.num_cols)
    bold = Font(bold=True)
    for r in range(1, header_rows + 1):
        for c in range(1, table.num_cols + 1):
            ws.cell(row=r, column=c).font = bold
    if header_rows or header_cols:
        # A coordinate, not ws.cell(): the anchor may fall inside a merged range,
        # and openpyxl cannot take a MergedCell here.
        ws.freeze_panes = f"{get_column_letter(header_cols + 1)}{header_rows + 1}"

    for c in range(table.num_cols):
        try:
            points = table.col_width(c)
        except Exception:  # noqa: BLE001 - width is cosmetic, never fail the conversion
            break
        if points:
            ws.column_dimensions[get_column_letter(c + 1)].width = round(
                max(points / POINTS_PER_CHARACTER, 2.0), 2
            )
    return written, truncated


def convert(path: Path, out_path: Path) -> Path:
    """Convert one Numbers document to an xlsx workbook at `out_path`."""
    from numbers_parser import Document

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    total = clipped = 0
    with numbers_suffixed(path) as source:
        tables = table_titles(Document(source))
        if not tables:
            raise SystemExit(f"ERROR: {path.name} contains no tables")
        for sheet, table, title in tables:
            if table.num_rows > MAX_ROWS or table.num_cols > MAX_COLS:
                raise SystemExit(
                    f"ERROR: {sheet.name}/{table.name} is {table.num_rows}x{table.num_cols}, "
                    f"beyond Excel's {MAX_ROWS}x{MAX_COLS} limit"
                )
            written, truncated = write_table(wb.create_sheet(title), table)
            total += written
            clipped += truncated
            print(f"[ok] {sheet.name!r}/{table.name!r} -> sheet {title!r} "
                  f"({table.num_rows} rows x {table.num_cols} cols, {written} non-empty cells)")
        sheet_count = len(tables)

    if clipped:
        print(f"[warn] {clipped} cell(s) held more than {MAX_CELL_TEXT} characters and were truncated")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[ok] wrote {out_path} ({sheet_count} sheet(s), {total} non-empty cells)")
    return out_path


def list_tables(path: Path) -> None:
    """Print the sheet/table inventory of a Numbers document without converting it."""
    from numbers_parser import Document

    print(f"{path}:")
    with numbers_suffixed(path) as source:
        for sheet, table, title in table_titles(Document(source)):
            print(f"  {sheet.name!r}/{table.name!r} -> sheet {title!r} "
                  f"({table.num_rows} rows x {table.num_cols} cols)")


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Convert Apple Numbers (.numbers) spreadsheets to Excel (.xlsx).",
    )
    parser.add_argument("inputs", nargs="+", help="one or more .numbers files")
    parser.add_argument("-o", "--output", help="output .xlsx path (single input only)")
    parser.add_argument("--outdir", help="directory for outputs (default: alongside each input)")
    parser.add_argument("--overwrite", action="store_true", help="replace an existing .xlsx")
    parser.add_argument("--list", action="store_true",
                        help="list the sheets and tables instead of converting")
    args = parser.parse_args(argv)

    if args.output and len(args.inputs) > 1:
        parser.error("--output takes a single input; use --outdir for several")
    if args.output and args.outdir:
        parser.error("--output and --outdir are mutually exclusive")

    try:
        import numbers_parser  # noqa: F401
    except ImportError:
        print("ERROR: numbers-parser is not installed. Run: pip install numbers-parser",
              file=sys.stderr)
        return 2

    from numbers_parser import NumbersError

    failures = 0
    for raw in args.inputs:
        path = Path(raw).expanduser()
        if not path.exists():
            print(f"ERROR: {path} not found", file=sys.stderr)
            failures += 1
            continue

        reason = reject_reason(classify(path), path)
        if reason:
            print(f"ERROR: {reason}", file=sys.stderr)
            failures += 1
            continue

        if args.list:
            list_tables(path)
            continue

        if args.output:
            out_path = Path(args.output).expanduser()
        else:
            out_dir = Path(args.outdir).expanduser() if args.outdir else path.parent
            out_path = out_dir / (path.stem + ".xlsx")
        if out_path.exists() and not args.overwrite:
            print(f"ERROR: {out_path} already exists (pass --overwrite to replace it)",
                  file=sys.stderr)
            failures += 1
            continue

        try:
            convert(path, out_path)
        except NumbersError as exc:
            print(f"ERROR: could not read {path.name}: {exc}", file=sys.stderr)
            failures += 1

    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
