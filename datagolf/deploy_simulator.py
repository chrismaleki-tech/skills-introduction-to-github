#!/usr/bin/env python3
"""
Deploy the StatCaddy member simulator for the current week, end to end.

This automates the manual weekly process (the "PGA stat caddy" checklist):
  1. Read the week's field + stats (from the workbook, refreshed by update_workbook.py).
  2. Log into WordPress (statcaddygolf.com) via the custom /golflogin URL.
  3. Set the published player set to exactly this week's field (publish field, draft rest);
     create records for field players not yet in the roster.
  4. Mark this week's tournament term active (create it if new), deactivate the others.
  5. Upload the field-stats CSV (Player Data) so every field player's stats are current.
  6. Regenerate player data, then build the pairwise simulation from the site's export and
     upload it (Simulation Data).
  7. VERIFY as the live endpoint (run_simulation on sample pairs) BEFORE enabling.
  8. Enable the simulator only if verification passes.

FAIL-SAFE: the simulator is set to Disabled at the start and only re-Enabled at the very
end after verification. Any error leaves it Disabled (members see the "updating" message)
rather than a half-built field — so a failed run never corrupts the live tool.

Env:
  DATAGOLF_KEY, WP_URL, WP_USERNAME, WP_PASSWORD   (required for live deploy)
  WP_PASSWORD        real WordPress login password for /golflogin (browser).
                     Application passwords (WP_APP_PASSWORD) work for REST only and
                     cannot sign into wp-admin — do not reuse them here.
  WORKBOOK           path to the latest workbook (default datagolf/workbooks/PGA_stat_caddy_latest.xlsx)
  TOUR               default pga
  STAT_WEIGHTS       optional JSON of the 8 model weights (defaults to the owner's general weights)
  DRY_RUN            "1" to build the field/stats/CSVs and skip all WordPress writes
"""

import argparse
import datetime
import itertools
import json
import os
import re
import sys
import tempfile
import unicodedata

import openpyxl
import requests
from scipy.stats import norm

DG = "https://feeds.datagolf.com"
# Owner's general stat weights + the calibrated probability scale (see reverse_engineer_model.py).
DEFAULT_WEIGHTS = {"ott": 1.2, "pts": 0.25, "app": 1.3, "putt": 1.0, "arg": 1.0, "fit": 3.0, "form": 1.5, "hist": 3.0}
K_SCALE = 0.1154
# Export-field -> weight key (the site's player export uses these names).
EXPORT_MAP = {"driving_distance": "ott", "driving_accuracy": "pts", "approach_accuracy": "app",
              "putting_skill": "putt", "ag": "arg", "sg": "fit", "current_form": "form", "course_history": "hist"}
REGEN_FIELD_KEY = "field_67f22c3f0d287"  # ACF "Regenerate Players Data" toggle
# A tour event runs four days, Thursday through Sunday. Data Golf's start_date is round one
# (a Thursday for every event on the schedule), so the final round is three days later.
DAYS_TO_FINAL_ROUND = 3
# Spelled out rather than strftime("%B") so the label cannot follow the runner's locale.
MONTH_NAMES = ("January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December")
# SiteGround's bot filter can answer /golflogin with a JS interstitial (/.well-known/sgcaptcha/)
# instead of the form. It clears itself once the challenge script runs, but datacentre IPs —
# GitHub's runners included — draw a slower variant that outlasts Playwright's default timeout.
LOGIN_FORM = "#user_login"
SG_CAPTCHA = "/sgcaptcha/"
LOGIN_ATTEMPTS = 3
LOGIN_FORM_TIMEOUT_MS = 60_000
# ACF renders the settings tab bar with JS after the page loads. Before it exists the only
# nodes carrying a tab's text are a hidden <label> and its hidden source anchor, so a plain
# text= click can pick one of those and wait out its timeout on something never made visible.
ACF_TAB_BUTTON = "a.acf-tab-button:visible"
ADMIN_TAB = "Admin Functions"
TAB_TIMEOUT_MS = 60_000
# Enough pages to hold a full season of tournament terms, and a stop so a list that always
# offers a "next page" cannot spin forever.
TERM_PAGE_LIMIT = 20
# Signed-out visitors are served from SiteGround's page cache, so the deploy is not finished
# until that cache has let go of the rebuild's "disabled" page.
PURGE_ITEM = "#wp-admin-bar-SG_CachePress_Supercacher_Purge a"
DISABLED_MARK = "Simulator Currently Disabled"
PUBLIC_UA = ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) "
             "Chrome/151.0.0.0 Safari/537.36")
PURGE_ATTEMPTS = 3
PURGE_CHECKS = 6
PURGE_CHECK_WAIT_MS = 5_000
# SiteGround answers a challenged request with 202 and a meta-refresh, not the page.
CHALLENGE_STATUS = 202
LIVE, DISABLED, UNAVAILABLE = "live", "disabled", "unavailable"


def log(msg):
    print(msg, flush=True)


def norm_name(n):
    n = str(n).strip().replace("ø", "o").replace("Ø", "O")
    if "," in n:
        last, first = [p.strip() for p in n.split(",", 1)]
        n = f"{first} {last}"
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKD", n).encode("ascii", "ignore").decode()).lower()


def dg_get(path, key, **params):
    r = requests.get(f"{DG}{path}", params={**params, "file_format": "json", "key": key}, timeout=30)
    r.raise_for_status()
    return r.json()


def event_start_date(key, tour, event):
    """Data Golf's round-one date for the named event, or None if it is not on the schedule."""
    for row in dg_get("/get-schedule", key, tour=tour).get("schedule", []):
        if norm_name(row.get("event_name") or "") == norm_name(event):
            return row.get("start_date")
    return None


def public_page_state(base, event):
    """How the simulator page looks to a signed-out visitor.

    LIVE         the page is being served and carries this week's event
    DISABLED     the rebuild's "currently disabled" page is still cached
    UNAVAILABLE  SiteGround answered with its bot challenge, or with a page that names no
                 event — either way the check has learnt nothing and must ask again

    Absence of the disabled notice is not evidence on its own: the challenge page carries no
    notice either, and reading that as success would let a stale cache through.
    """
    r = requests.get(f"{base.rstrip('/')}/matchup-simulator/", timeout=30,
                     headers={"User-Agent": PUBLIC_UA})
    if r.status_code == CHALLENGE_STATUS or "SG-Captcha" in r.headers or SG_CAPTCHA in r.text:
        return UNAVAILABLE
    if DISABLED_MARK.lower() in r.text.lower():
        return DISABLED
    return LIVE if event.lower() in r.text.lower() else UNAVAILABLE


def format_event_dates(start_date):
    """The label the simulator shows beneath the tournament name, e.g. 'August 20-23'.

    A range that crosses a month gets the second month spelled out too ('April 30 - May 3').
    """
    start = datetime.date.fromisoformat(start_date)
    end = start + datetime.timedelta(days=DAYS_TO_FINAL_ROUND)
    if start.month == end.month:
        return f"{MONTH_NAMES[start.month - 1]} {start.day}-{end.day}"
    return f"{MONTH_NAMES[start.month - 1]} {start.day} - {MONTH_NAMES[end.month - 1]} {end.day}"


# ── Build the week's field + stats ──────────────────────────────────────────

def build_field(key, workbook, tour):
    """Return (event_name, [ {name, ott, pts, app, putt, arg, fit, form, hist} ]).

    Field = Data Golf's current entrants. Stats come from the workbook's PGA Database
    where the player is tracked (real Form/DG Points), else from Data Golf (skill-ratings
    + decompositions, Form≈sg_total, DG Points≈field median). Players Data Golf cannot
    rate at all (no skill-ratings) are skipped, so no fabricated strokes-gained ship.
    """
    fu = dg_get("/field-updates", key, tour=tour)
    event = fu.get("event_name")
    field = [{"dg_id": p["dg_id"], "name": p["player_name"]} for p in fu.get("field", [])]
    log(f"Data Golf field: {event} — {len(field)} entrants")

    skill = {p["dg_id"]: p for p in dg_get("/preds/skill-ratings", key, display="value").get("players", [])}
    dec = {p["dg_id"]: p for p in dg_get("/preds/player-decompositions", key, tour=tour).get("players", [])}

    wb_stats = {}
    if workbook and os.path.exists(workbook):
        try:
            db = openpyxl.load_workbook(workbook, data_only=True)["PGA Database"]
            hdr = {str(c.value).strip(): i + 1 for i, c in enumerate(db[1]) if c.value}
            cols = {"ott": "SG OTT", "pts": "Points", "app": "Approach", "putt": "Putting",
                    "arg": "Around Green", "fit": "T2Green", "form": "Form", "hist": "History"}
            for r in range(2, db.max_row + 1):
                nm = db.cell(row=r, column=1).value
                if not nm:
                    continue
                rec = {}
                ok = True
                for k, h in cols.items():
                    v = db.cell(row=r, column=hdr[h]).value if h in hdr else None
                    try:
                        rec[k] = round(float(v), 4)
                    except (TypeError, ValueError):
                        ok = False
                        break
                if ok:
                    wb_stats[norm_name(nm)] = rec
            log(f"workbook PGA Database: {len(wb_stats)} players with complete stats")
        except Exception as exc:  # noqa: BLE001 - workbook optional
            log(f"[warn] could not read workbook ({exc}); using Data Golf stats only")

    # neutral DG-points fill = median of available points
    pts_pool = [v["pts"] for v in wb_stats.values()] or [0.0]
    med_pts = round(sorted(pts_pool)[len(pts_pool) // 2], 3)

    players, skipped = [], []
    for p in field:
        k = norm_name(p["name"])
        first_last = p["name"]
        if "," in first_last:
            last, first = [x.strip() for x in first_last.split(",", 1)]
            first_last = f"{first} {last}"
        if k in wb_stats:
            rec = dict(wb_stats[k])
        elif p["dg_id"] in skill:
            s = skill[p["dg_id"]]
            d = dec.get(p["dg_id"], {})
            rec = {"ott": round(s.get("sg_ott", 0), 4), "pts": med_pts, "app": round(s.get("sg_app", 0), 4),
                   "putt": round(s.get("sg_putt", 0), 4), "arg": round(s.get("sg_arg", 0), 4),
                   "fit": round(d.get("total_fit_adjustment", 0), 2), "form": round(s.get("sg_total", 0), 2),
                   "hist": round(d.get("total_course_history_adjustment", 0), 2)}
        else:
            skipped.append(first_last)
            continue
        rec["name"] = first_last
        players.append(rec)
    if skipped:
        log(f"[note] {len(skipped)} entrant(s) unrated by Data Golf — excluded: {', '.join(sorted(skipped))}")
    log(f"simulatable field: {len(players)} players")
    return event, players


FIELD_HEADER = ["Player", "SG OTT", "Points", "Approach", "Putting", "Around Green", "T2Green", "Form", "History", "Tournament", "Tour"]


def _field_rows(event, players):
    return [[p["name"], p["ott"], p["pts"], p["app"], p["putt"], p["arg"], p["fit"], p["form"], p["hist"], event, "PGA Tour"] for p in players]


def write_field_csv(event, players, path):
    import csv
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(FIELD_HEADER)
        w.writerows(_field_rows(event, players))


def write_field_sheet(event, players, out_dir):
    """Write the week's downloadable field-stat sheet as CSV + XLSX (a stable 'latest' plus
    an event/date-stamped copy) so it can be grabbed each week from the repo."""
    import csv
    from datetime import date
    os.makedirs(out_dir, exist_ok=True)
    slug = re.sub(r"[^A-Za-z0-9]+", "_", event).strip("_")
    stamp = date.today().isoformat()
    rows = _field_rows(event, players)
    targets = [
        os.path.join(out_dir, "StatCaddy_Field_latest.csv"),
        os.path.join(out_dir, f"StatCaddy_{slug}_Field_{stamp}.csv"),
    ]
    for t in targets:
        with open(t, "w", newline="") as f:
            w = csv.writer(f); w.writerow(FIELD_HEADER); w.writerows(rows)
    for xlsx in [os.path.join(out_dir, "StatCaddy_Field_latest.xlsx"),
                 os.path.join(out_dir, f"StatCaddy_{slug}_Field_{stamp}.xlsx")]:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Field"
        ws.append(FIELD_HEADER)
        for r in rows:
            ws.append(r)
        wb.save(xlsx)
    log(f"field-stat sheet written: {len(players)} players -> {out_dir}/StatCaddy_Field_latest.(csv|xlsx) (+ dated)")


# ── WordPress deployment (browser-driven) ─────────────────────────────────────

class Deployer:
    def __init__(self, pg, base):
        self.pg = pg
        self.base = base.rstrip("/")
        self.SET = f"{self.base}/wp-admin/edit.php?post_type=player&page=additional-settings"

    def open_login(self):
        """Land on the real login form, sitting out SiteGround's challenge if one is served."""
        last = None
        for attempt in range(1, LOGIN_ATTEMPTS + 1):
            self.pg.goto(f"{self.base}/golflogin", wait_until="domcontentloaded")
            try:
                self.pg.wait_for_selector(LOGIN_FORM, timeout=LOGIN_FORM_TIMEOUT_MS)
                return
            except Exception as exc:  # noqa: BLE001
                last = exc
                held = SG_CAPTCHA in self.pg.url
                log(f"  login form did not appear (attempt {attempt}/{LOGIN_ATTEMPTS})"
                    + (" — held at the SiteGround challenge" if held else ""))
                self.pg.wait_for_timeout(5_000)
        raise SystemExit(f"could not reach the login form at {self.base}/golflogin: {last}")

    def login(self, user, pw):
        self.open_login()
        self.pg.fill("#user_login", user)
        self.pg.fill("#user_pass", pw)
        self.pg.click("#wp-submit")
        self.pg.wait_for_load_state("networkidle")
        if "/wp-admin" not in self.pg.url:
            raise SystemExit("login failed (not redirected to wp-admin)")
        # raise per-page so bulk lists fit one page
        self.pg.goto(f"{self.base}/wp-admin/edit.php?post_type=player&all_posts=1", wait_until="networkidle")
        try:
            self.pg.click("#show-settings-link"); self.pg.wait_for_timeout(400)
            self.pg.fill("input.screen-per-page", "600"); self.pg.click("#screen-options-apply")
            self.pg.wait_for_load_state("networkidle")
        except Exception:
            pass

    def _admin_functions(self):
        self.pg.goto(self.SET, wait_until="networkidle")
        tab = self.pg.locator(ACF_TAB_BUTTON).filter(has_text=ADMIN_TAB).first
        tab.wait_for(state="visible", timeout=TAB_TIMEOUT_MS)
        tab.click()
        self.pg.wait_for_timeout(900)

    def set_status(self, value):
        self._admin_functions()
        self.pg.select_option("select", value)
        self.pg.locator("#publish").scroll_into_view_if_needed()
        self.pg.click("#publish", force=True)
        self.pg.wait_for_load_state("networkidle"); self.pg.wait_for_timeout(1500)

    def roster(self):
        self.pg.goto(f"{self.base}/wp-admin/edit.php?post_type=player&all_posts=1", wait_until="networkidle")
        return self.pg.evaluate("""()=>{const o=[];document.querySelectorAll('tbody#the-list tr').forEach(tr=>{const cb=tr.querySelector('input[name="post[]"]');const t=tr.querySelector('.row-title');if(cb&&t)o.push({id:cb.value,name:t.textContent.trim()});});return o;}""")

    def create_player(self, name):
        self.pg.goto(f"{self.base}/wp-admin/post-new.php?post_type=player", wait_until="networkidle")
        self.pg.wait_for_timeout(1200)
        self.pg.fill("#title", name); self.pg.wait_for_timeout(400)
        self.pg.click("#publish"); self.pg.wait_for_load_state("networkidle"); self.pg.wait_for_timeout(1500)

    def bulk_status(self, ids, status):
        if not ids:
            return
        self.pg.goto(f"{self.base}/wp-admin/edit.php?post_type=player&all_posts=1", wait_until="networkidle")
        for pid in ids:
            cb = self.pg.locator(f"#cb-select-{pid}")
            if cb.count():
                cb.check()
        self.pg.select_option("#bulk-action-selector-top", "edit")
        self.pg.click("#doaction"); self.pg.wait_for_timeout(1400)
        self.pg.select_option('select[name="_status"]', status)
        self.pg.click("#bulk_edit"); self.pg.wait_for_load_state("networkidle"); self.pg.wait_for_timeout(2500)

    def fill_term_dates(self, dates):
        """Put `dates` in the term's Tournament Dates box. True when the value changed.

        The box is free text that nobody re-types once the season moves on, which is how the
        simulator ended up advertising last year's dates for most of the calendar.
        """
        box = self.pg.locator('.acf-field[data-name="tournament_dates"] input').first
        if not box.count():
            log("  [warn] term has no Tournament Dates field — leaving the label alone")
            return False
        if (box.input_value() or "").strip() == dates:
            return False
        box.fill(dates)
        return True

    def term_rows(self):
        """Every tournament term, following the admin list's pagination.

        A full season is more terms than WordPress shows on one screen, and the events that
        fall off page one are exactly the late-season ones — the TOUR Championship among them.
        Reading only the first screen left those terms unvisited, so they kept whatever dates
        they were last given by hand.
        """
        rows, page = [], 1
        while page <= TERM_PAGE_LIMIT:
            self.pg.goto(f"{self.base}/wp-admin/edit-tags.php?taxonomy=tournament&post_type=player&paged={page}",
                         wait_until="networkidle")
            found = self.pg.evaluate("""()=>{const o=[];document.querySelectorAll('a.row-title').forEach(a=>{const m=(a.getAttribute('href')||'').match(/tag_ID=(\\d+)/);if(m)o.push({id:m[1],name:a.textContent.trim()});});return o;}""")
            if not found:
                break
            rows += found
            if not self.pg.locator("a.next-page").count():
                break
            page += 1
        return rows

    def set_active_tournament(self, event, dates=None):
        rows = self.term_rows()
        if not any(norm_name(r["name"]) == norm_name(event) for r in rows):
            self.pg.goto(f"{self.base}/wp-admin/edit-tags.php?taxonomy=tournament&post_type=player", wait_until="networkidle")
            self.pg.fill("#tag-name", event)
            self.pg.click("#submit"); self.pg.wait_for_load_state("networkidle"); self.pg.wait_for_timeout(1500)
            rows = self.term_rows()
        log(f"tournament terms found: {len(rows)}")
        # deactivate any active, activate ours, and date the one we activate
        for r in rows:
            self.pg.goto(f"{self.base}/wp-admin/term.php?taxonomy=tournament&post_type=player&tag_ID={r['id']}", wait_until="networkidle")
            cb = self.pg.locator("input.acf-switch-input").first
            if not cb.count():
                continue
            want = (norm_name(r["name"]) == norm_name(event))
            dirty = False
            if want and dates:
                dirty = self.fill_term_dates(dates)
                if dirty:
                    log(f"  dates for {r['name']}: {dates}")
            if cb.is_checked() != want:
                self.pg.evaluate("([w])=>{const c=document.querySelector('input.acf-switch-input'); c.checked=w; c.dispatchEvent(new Event('change',{bubbles:true}));}", [want])
                dirty = True
            if dirty:
                self.pg.locator('input#submit, input[type="submit"][value="Update"]').first.click()
                self.pg.wait_for_load_state("networkidle"); self.pg.wait_for_timeout(800)

    def upload_csv(self, which, path):
        """which: 'player' (Player Data, last Add File) or 'sim' (Simulation Data, first)."""
        self._admin_functions()
        idx = self.pg.locator("a:has-text('Add File'), button:has-text('Add File')")
        (idx.last if which == "player" else idx.first).click()
        self.pg.wait_for_timeout(1500)
        self.pg.locator("input[type=file]").last.set_input_files(path)
        self.pg.wait_for_timeout(5000)
        for sel in [".media-button-select", "button:has-text('Select')", ".media-modal button.button-primary"]:
            loc = self.pg.locator(sel)
            if loc.count() and loc.last.is_visible():
                loc.last.click(); break
        self.pg.wait_for_timeout(1500)
        self.pg.locator("#publish").scroll_into_view_if_needed()
        self.pg.click("#publish", force=True)
        self.pg.wait_for_load_state("networkidle"); self.pg.wait_for_timeout(3000)

    def regenerate(self):
        self._admin_functions()
        self.pg.evaluate(f"()=>{{const c=document.querySelector('#acf-{REGEN_FIELD_KEY}'); if(c&&!c.checked){{c.checked=true;c.dispatchEvent(new Event('change',{{bubbles:true}}));}}}}")
        self.pg.evaluate("document.querySelector('#publish').click()")
        self.pg.wait_for_load_state("networkidle"); self.pg.wait_for_timeout(4000)

    def read_export(self):
        self._admin_functions()
        ta = self.pg.locator("textarea").filter(has_text="players = [").first
        return json.loads(re.search(r"\[.*\]", ta.input_value(), re.S).group(0))

    def purge_cache(self):
        """Purge SiteGround's page cache. True when a purge was actually triggered.

        The admin-bar entry lives in a hover submenu, so follow its href instead of clicking
        it. This used to look for a node id the plugin does not use and skip the purge in
        silence, which left signed-out members on the rebuild's "updating" page for as long
        as the cache entry lived — a successful deploy that nobody could see.
        """
        self.pg.goto(f"{self.base}/wp-admin/", wait_until="networkidle")
        link = self.pg.locator(PURGE_ITEM).first
        if not link.count():
            log("  [warn] SiteGround purge control not found — signed-out visitors may be served stale HTML")
            return False
        self.pg.goto(link.get_attribute("href"), wait_until="domcontentloaded")
        self.pg.wait_for_timeout(2500)
        return True

    def settle_public_page(self, event):
        """Purge until a signed-out request is served this week's event."""
        for attempt in range(1, PURGE_ATTEMPTS + 1):
            self.purge_cache()
            state = UNAVAILABLE
            for _ in range(PURGE_CHECKS):
                self.pg.wait_for_timeout(PURGE_CHECK_WAIT_MS)
                state = public_page_state(self.base, event)
                if state == LIVE:
                    return True
            log(f"  signed-out page reads '{state}' after purge {attempt}/{PURGE_ATTEMPTS}")
        return False

    def verify(self, tournament_slug, sample_rows):
        """Run run_simulation on sample pairs; return True if they match the sim file."""
        self.pg.goto(f"{self.base}/matchup-simulator/", wait_until="networkidle")
        self.pg.wait_for_timeout(1500)
        rs = re.search(r'id="run_simulation_nonce"[^>]*value="([a-f0-9]+)"', self.pg.content()).group(1)
        js = """async([n,a,bb])=>new Promise(r=>{jQuery.ajax({url:siteData.adminAjax,type:'post',dataType:'json',data:{action:'run_simulation',players:[[a],[bb]],user:siteData.currentUser,nonce:n},success:d=>r(d),error:x=>r(null)});});"""
        ok = 0
        for a, b, expected in sample_rows:
            res = None
            for _ in range(4):
                res = self.pg.evaluate(js, [rs, a, b])
                if isinstance(res, list):
                    break
                self.pg.wait_for_timeout(1200)
            live = next((x["win_percent"] for x in res if str(x["player_id"]) == a), None) if isinstance(res, list) else None
            # Tolerance sits below the displayed tenth: the live value should be the file's
            # value, so anything the simulator re-rounds differently has to fail here.
            match = live is not None and abs(float(live) - float(expected)) < 0.05
            if match:
                ok += 1
            log(f"  verify {a} vs {b}: live={live} expected={expected}{'' if match else '  <-- MISMATCH'}")
        return ok == len(sample_rows)


def build_sim_csv(export, weights, path):
    """Write every ordered pairing as `id_a,id_b,win_pct,wins_of_10000`.

    The simulator displays `wins_of_10000 / 100` rounded to one decimal, so counts are put
    on a tenth-of-a-percent grid and the opposite side is derived by subtraction rather than
    recomputed. Off the grid, a count ending in 5 makes both halves of a matchup display a
    trailing …x5, PHP's round() lifts both away from zero, and the two sides add up to 100.1.
    """
    R = {}
    for p in export:
        R[p["id"]] = sum(weights[EXPORT_MAP[f]] * float(p[f]) for f in EXPORT_MAP)
    ids = [p["id"] for p in export]
    counts = {}
    for a, b in itertools.combinations(ids, 2):
        cnt = int(round(norm.cdf(K_SCALE * (R[a] - R[b])) * 1000)) * 10
        counts[(a, b)], counts[(b, a)] = cnt, 10000 - cnt
    off = [(a, b) for (a, b), c in counts.items() if c % 10 or c + counts[(b, a)] != 10000]
    if off:
        raise ValueError(f"{len(off)} pairing(s) would not display a total of 100.0, e.g. {off[0]}")
    rows = {}
    with open(path, "w", newline="") as f:
        for a, b in itertools.permutations(ids, 2):
            cnt = counts[(a, b)]
            f.write(f"{a},{b},{cnt / 100:.1f},{cnt}\n")
            rows[(str(a), str(b))] = cnt / 100
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=os.path.join(os.path.dirname(__file__), "workbooks", "PGA_stat_caddy_latest.xlsx"))
    ap.add_argument("--tour", default=os.environ.get("TOUR", "pga"))
    ap.add_argument("--out-dir", default="/tmp")
    ap.add_argument("--field-dir", default=os.path.join(os.path.dirname(__file__), "workbooks"),
                    help="Where to save the downloadable weekly field-stat sheet (CSV + XLSX).")
    args = ap.parse_args()

    key = os.environ.get("DATAGOLF_KEY")
    if not key:
        log("ERROR: DATAGOLF_KEY not set"); return 2
    weights = json.loads(os.environ["STAT_WEIGHTS"]) if os.environ.get("STAT_WEIGHTS") else DEFAULT_WEIGHTS
    dry = os.environ.get("DRY_RUN") == "1"

    event, players = build_field(key, args.workbook, args.tour)
    if len(players) < 2:
        log("ERROR: fewer than 2 simulatable players — aborting"); return 3
    field_csv = os.path.join(args.out_dir, "deploy_field.csv")
    write_field_csv(event, players, field_csv)
    # always publish the downloadable weekly field-stat sheet (CSV + XLSX)
    write_field_sheet(event, players, args.field_dir)

    start = event_start_date(key, args.tour, event)
    event_dates = format_event_dates(start) if start else None
    if event_dates:
        log(f"event dates from the Data Golf schedule: {event_dates} (round one {start})")
    else:
        log(f"[warn] '{event}' is not on the Data Golf schedule — leaving its dates as they are")

    if dry:
        log(f"[dry-run] built {len(players)} players for '{event}' ({event_dates}). Skipping WordPress writes.")
        return 0

    for v in ("WP_URL", "WP_USERNAME", "WP_PASSWORD"):
        if not os.environ.get(v):
            log(f"ERROR: {v} not set"); return 2
    if os.environ.get("WP_APP_PASSWORD") and os.environ["WP_PASSWORD"] == os.environ["WP_APP_PASSWORD"]:
        log("ERROR: WP_PASSWORD looks like WP_APP_PASSWORD — browser login needs the real account password")
        return 2

    from playwright.sync_api import sync_playwright
    field_norm = {norm_name(p["name"]) for p in players}
    slug = re.sub(r"[^a-z0-9]+", "-", event.lower()).strip("-")

    with sync_playwright() as pw:
        b = pw.chromium.launch()
        ctx = b.new_context(viewport={"width": 1500, "height": 2000})
        pg = ctx.new_page()
        d = Deployer(pg, os.environ["WP_URL"])
        try:
            d.login(os.environ["WP_USERNAME"], os.environ["WP_PASSWORD"])
            d.set_status("disabled")  # FAIL-SAFE: down for the duration of the rebuild

            roster = d.roster()
            by_norm = {}
            for r in roster:
                by_norm.setdefault(norm_name(r["name"]), r["id"])
            # create any field player missing from the roster
            for p in players:
                if norm_name(p["name"]) not in by_norm:
                    log(f"creating new player record: {p['name']}")
                    d.create_player(p["name"])
            roster = d.roster()
            by_norm = {}
            for r in roster:
                by_norm.setdefault(norm_name(r["name"]), r["id"])

            publish_ids = [by_norm[k] for k in field_norm if k in by_norm]
            draft_ids = [r["id"] for r in roster if norm_name(r["name"]) not in field_norm]
            log(f"publishing {len(publish_ids)} field players; drafting {len(draft_ids)} others")
            d.bulk_status(publish_ids, "publish")
            d.bulk_status(draft_ids, "draft")

            d.set_active_tournament(event, event_dates)
            d.upload_csv("player", field_csv)     # set stats
            d.regenerate()

            export = d.read_export()
            log(f"export after regenerate: {len(export)} players")
            sim_csv = os.path.join(args.out_dir, "deploy_sim.csv")
            rows = build_sim_csv(export, weights, sim_csv)
            d.upload_csv("sim", sim_csv)
            d.purge_cache()

            import random
            sample = [(a, b, rows[(a, b)]) for a, b in random.sample(list(rows), min(3, len(rows)))]
            if d.verify(slug, sample):
                d.set_status("enabled")
                if not d.settle_public_page(event):
                    log(f"ERROR: signed-out visitors are not being served '{event}'. The "
                        "simulator is enabled — purge SiteGround's cache by hand.")
                    return 6
                log(f"SUCCESS: '{event}' live with {len(export)} players.")
                return 0
            log("ERROR: verification failed — leaving simulator DISABLED (safe state).")
            return 4
        except Exception as exc:  # noqa: BLE001
            log(f"ERROR during deploy: {exc}. Leaving simulator DISABLED (safe state).")
            try:
                d.set_status("disabled")
            except Exception:
                pass
            return 5
        finally:
            b.close()


if __name__ == "__main__":
    sys.exit(main())
