#!/usr/bin/env python3
"""
Regenerate the public "Build Your Own Matchup Model" simulator and (optionally)
push it to the live WordPress page.

Data source preference:
  1. --field-csv / StatCaddy_Field_latest.csv (current event field sheet)
  2. Workbook PGA Database rows with a T2Green (course-fit) value

Each player's widget stats are read, an embeddable WordPress-safe HTML block is
built (scoped CSS + base64 JS + inline JSON), and — if WordPress credentials are
present — the block is written to the simulator page via REST.

Env (only needed for the push step):
    WP_URL, WP_USERNAME, WP_APP_PASSWORD
    WP_SIMULATOR_PAGE_ID   (defaults to 4952)

Usage:
    python datagolf/build_public_simulator.py --workbook datagolf/workbooks/PGA_stat_caddy_latest.xlsx
    python datagolf/build_public_simulator.py --push
"""

import argparse
import base64
import csv
import json
import os
import re
import sys

import openpyxl
import requests

REPO_DIR = os.path.dirname(__file__)
SIM_DIR = os.path.join(os.path.dirname(REPO_DIR), "simulator")
TEMPLATE = os.path.join(SIM_DIR, "statcaddy-simulator-standalone.html")
DEFAULT_FIELD_CSV = os.path.join(REPO_DIR, "workbooks", "StatCaddy_Field_latest.csv")

# Field / PGA Database column header -> widget stat key
COLMAP = {
    "SG OTT": "ott", "Approach": "app", "Around Green": "arg", "Putting": "putt",
    "Form": "form", "History": "hist", "Points": "pts",
}

# Mobile chrome that lives outside #sc-sim (Elementor header + simulator switcher + 1v1 cards).
PAGE_CHROME_CSS = (
    " .page-id-4952 .page-header,.page-id-4952 .entry-title{display:none!important}"
    " .page-id-4952 .elementor-location-header .elementor-element-080483f,"
    ".page-id-4952 .elementor-location-header .elementor-element-d094220{position:relative!important}"
    " .page-id-4952 .sc-header-matchup-title{position:absolute;left:50%;top:50%;transform:translate(-50%,-50%);"
    "margin:0;padding:0 4px;font-family:Staatliches,Sans-serif;font-size:clamp(14px,4.2vw,28px);"
    "font-weight:400;letter-spacing:1px;line-height:1.05;color:#1b4332;text-transform:uppercase;"
    "white-space:normal;max-width:min(54vw,260px);text-align:center;pointer-events:none;z-index:5;"
    "overflow-wrap:anywhere;hyphens:auto}"
    " @media (max-width:782px){"
    ".page-id-4952 .sc-more-sims,.page-id-166 .sc-more-sims,.page-id-3736 .sc-more-sims,"
    ".page-id-1738 .sc-more-sims{margin-top:64px!important;margin-bottom:10px!important}"
    # Beat theme #player-comparisons .players-outer-container{max-width:45%!important}
    ".page-id-166 #player-comparisons .player-comparison-container.row,"
    ".page-id-3736 #player-comparisons-multiple .player-comparison-container.row{"
    "display:flex!important;flex-direction:column!important;flex-wrap:nowrap!important;"
    "align-items:stretch!important;gap:18px!important}"
    ".page-id-166 #player-comparisons #player1-outer.players-outer-container,"
    ".page-id-166 #player-comparisons #player2-outer.players-outer-container,"
    ".page-id-166 #player-comparisons .players-outer-container,"
    ".page-id-3736 #player-comparisons-multiple .players-outer-container{"
    "flex:0 0 auto!important;max-width:100%!important;width:100%!important;float:none!important}"
    ".page-id-166 #player-comparisons .player-select,"
    ".page-id-3736 #player-comparisons-multiple .player-select,"
    ".page-id-166 #player-comparisons select,"
    ".page-id-3736 #player-comparisons-multiple select,"
    ".page-id-166 #player-comparisons .select2-container,"
    ".page-id-3736 #player-comparisons-multiple .select2-container{"
    "max-width:100%!important;width:100%!important;box-sizing:border-box!important;font-size:16px!important}"
    ".page-id-166 #player-comparisons button,.page-id-3736 #player-comparisons-multiple button{"
    "max-width:100%!important;box-sizing:border-box!important;white-space:normal!important;"
    "min-height:44px!important}"
    "}"
)


def _record_from_mapping(name, getter):
    rec = {"name": str(name).strip()}
    for header, key in COLMAP.items():
        v = getter(header)
        try:
            rec[key] = round(float(v), 3)
        except (TypeError, ValueError):
            return None
    return rec


def build_players_from_csv(path: str) -> list:
    players = []
    with open(path, newline="") as f:
        for row in csv.DictReader(f):
            name = row.get("Player") or row.get("player")
            if not name:
                continue
            rec = _record_from_mapping(name, lambda h, r=row: r.get(h))
            if rec:
                players.append(rec)
    players.sort(key=lambda p: p["name"])
    return players


def build_players_from_workbook(workbook: str) -> list:
    wb = openpyxl.load_workbook(workbook, data_only=True)
    db = wb["PGA Database"]
    hdr = {str(c.value).strip(): i + 1 for i, c in enumerate(db[1]) if c.value}
    fit_col = hdr.get("T2Green")
    players = []
    for r in range(2, db.max_row + 1):
        name = db.cell(row=r, column=1).value
        if not name:
            continue
        if fit_col and db.cell(row=r, column=fit_col).value is None:
            continue
        rec = _record_from_mapping(
            name,
            lambda h, row=r: db.cell(row=row, column=hdr[h]).value if h in hdr else None,
        )
        if rec:
            players.append(rec)
    players.sort(key=lambda p: p["name"])
    return players


def build_players(workbook: str, field_csv: str | None = None) -> list:
    csv_path = field_csv if field_csv is not None else DEFAULT_FIELD_CSV
    if csv_path and os.path.isfile(csv_path):
        players = build_players_from_csv(csv_path)
        print(f"[ok] field source: {csv_path} ({len(players)} players)")
        return players
    print(f"[warn] field csv missing — falling back to workbook T2Green field")
    return build_players_from_workbook(workbook)


def build_embed(players: list) -> str:
    html = open(TEMPLATE).read()
    style = re.sub(
        r"\s+", " ",
        re.search(r"<style>(.*?)</style>", html, re.S).group(1).replace("body {", "#sc-sim {"),
    )
    style += PAGE_CHROME_CSS
    body = re.search(r"<body>(.*?)</body>", html, re.S).group(1)
    js = re.search(r"<script>(.*?)</script>", body, re.S).group(1)
    markup = re.sub(r"\n\s*\n+", "\n", re.sub(r"<script.*?</script>", "", body, flags=re.S))
    # Outer #sc-sim already exists in template markup — unwrap so we don't nest wrappers.
    markup = re.sub(r'^\s*<div id="sc-sim">', "", markup)
    markup = re.sub(r"</div>\s*$", "", markup.strip())
    data = json.dumps(players, separators=(",", ":"))
    return (
        f'<div id="sc-sim"><style>{style}</style>{markup}'
        f'<script type="application/json" id="embedded-data">{data}</script>'
        f'<script>eval(atob("{base64.b64encode(js.encode()).decode()}"));</script></div>'
    )


def push_to_wordpress(embed: str) -> None:
    url = os.environ["WP_URL"].rstrip("/")
    page_id = os.environ.get("WP_SIMULATOR_PAGE_ID", "4952")
    auth = (os.environ["WP_USERNAME"], os.environ["WP_APP_PASSWORD"].replace(" ", ""))
    r = requests.post(
        f"{url}/wp-json/wp/v2/pages/{page_id}",
        auth=auth,
        json={"content": embed},
        timeout=60,
    )
    r.raise_for_status()
    rendered = r.json()["content"]["rendered"]
    n = len(json.loads(re.search(r'id="embedded-data">(.*?)</script>', rendered, re.S).group(1)))
    print(f"[ok] pushed to WordPress page {page_id}: {n} players live")


def main() -> int:
    parser = argparse.ArgumentParser(description="Rebuild the public matchup simulator.")
    parser.add_argument("--workbook", default=os.path.join(REPO_DIR, "workbooks", "PGA_stat_caddy_latest.xlsx"))
    parser.add_argument("--field-csv", default=DEFAULT_FIELD_CSV,
                        help="Current-event field CSV (default: StatCaddy_Field_latest.csv)")
    parser.add_argument("--push", action="store_true", help="Also update the live WordPress page.")
    args = parser.parse_args()

    players = build_players(args.workbook, args.field_csv)
    if len(players) < 2:
        print(f"ERROR: only {len(players)} field players found — aborting.", file=sys.stderr)
        return 1
    embed = build_embed(players)
    json.dump(players, open(os.path.join(SIM_DIR, "players.json"), "w"), indent=0)
    open(os.path.join(SIM_DIR, "wp-embed.html"), "w").write(embed)
    print(
        f"[ok] rebuilt simulator with {len(players)} field players "
        f"(default: {players[0]['name']} vs {players[1]['name']})"
    )

    if args.push:
        if not all(os.environ.get(v) for v in ("WP_URL", "WP_USERNAME", "WP_APP_PASSWORD")):
            print("ERROR: --push requires WP_URL / WP_USERNAME / WP_APP_PASSWORD", file=sys.stderr)
            return 2
        push_to_wordpress(embed)
    return 0


if __name__ == "__main__":
    sys.exit(main())
