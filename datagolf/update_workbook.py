#!/usr/bin/env python3
"""
Update the "PGA stat caddy" workbook with new tournament weeks from Data Golf.

The workbook tracks rolling windows of per-round strokes-gained per player:
  - "Current Form"  : SG total, last 10 weeks, Form = TRIMMEAN(B:K, 2/COUNT)
  - "SG OTT/APP/ARG/PUTT" : SG components, last 15 weeks, same TRIMMEAN pattern
  - "DG Points"     : requires Data Golf's historical tier; left blank for new weeks

For each new week this script shifts every sheet's date window left (dropping the
oldest week) and writes the new week's values in the last date column, so the
TRIMMEAN formulas and the PGA Database cross-references stay untouched.

Week data sources:
  --live            : current event's final stats from /preds/live-tournament-stats
                      (round=event_avg -> already per-round)
  --snapshot DIR    : a committed pull_all.py snapshot (dg_live_tournament_stats.csv
                      cumulative SG divided by rounds played from dg_in_play.csv)

Usage:
    python datagolf/update_workbook.py workbook.xlsx out.xlsx \
        --snapshot datagolf/data --snapshot-date 2026-07-12 \
        --live --live-date 2026-07-19
"""

import argparse
import os
import re
import sys
import unicodedata
from datetime import datetime

import openpyxl
import pandas as pd
import requests

STAT_SHEETS = {
    "SG OTT": "sg_ott",
    "SG APP": "sg_app",
    "SG ARG": "sg_arg",
    "SG PUTT": "sg_putt",
    "Current Form": "sg_total",
    "Class": None,  # DG Points data (sheet labelled "Class"); window shifts, values from archive
}


def norm_name(name: str) -> str:
    """Normalize 'Ludvig Aberg  ' / 'Aberg, Ludvig' to a comparable key."""
    name = str(name).strip()
    if "," in name:
        last, first = [p.strip() for p in name.split(",", 1)]
        name = f"{first} {last}"
    name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", name).lower()


def week_from_live(api_key: str) -> dict:
    """Per-round SG averages for the current event, keyed by normalized name."""
    resp = requests.get(
        "https://feeds.datagolf.com/preds/live-tournament-stats",
        params={
            "stats": "sg_ott,sg_app,sg_arg,sg_putt,sg_total",
            "round": "event_avg",
            "display": "value",
            "file_format": "json",
            "key": api_key,
        },
        timeout=30,
    )
    resp.raise_for_status()
    payload = resp.json()
    print(f"live event: {payload.get('event_name')} (updated {payload.get('last_updated')})")
    out = {}
    for p in payload.get("live_stats") or []:
        out[norm_name(p["player_name"])] = {
            stat: p.get(stat) for stat in ("sg_ott", "sg_app", "sg_arg", "sg_putt", "sg_total")
        }
    return out


def week_from_snapshot(snapshot_dir: str) -> dict:
    """Per-round SG from a pull_all.py snapshot (cumulative / rounds played)."""
    stats = pd.read_csv(os.path.join(snapshot_dir, "dg_live_tournament_stats.csv"))
    inplay = pd.read_csv(os.path.join(snapshot_dir, "dg_in_play.csv"))
    print(f"snapshot event: {stats['event_name'].iloc[0]}")
    rounds = inplay.set_index("dg_id")[["R1", "R2", "R3", "R4"]].notna().sum(axis=1)
    out = {}
    for _, row in stats.iterrows():
        n = int(rounds.get(row["dg_id"], 0)) or None
        if n is None:
            continue
        out[norm_name(row["player_name"])] = {
            stat: (row[stat] / n if pd.notna(row.get(stat)) else None)
            for stat in ("sg_ott", "sg_app", "sg_arg", "sg_putt", "sg_total")
        }
    return out


def week_from_archive(api_key: str, week: str) -> dict:
    """Official per-round SG averages for every event completing on `week` (Sunday).

    Merges all events that week (main + opposite field). Preferred over live capture
    once Data Golf archives the event, since archived SG reflects their final revisions.
    """
    listing = requests.get(
        "https://feeds.datagolf.com/historical-event-data/event-list",
        params={"tour": "pga", "file_format": "json", "key": api_key}, timeout=30,
    ).json()
    out = {}
    for ev in listing:
        if ev.get("date") != week:
            continue
        detail = requests.get(
            "https://feeds.datagolf.com/historical-raw-data/rounds",
            params={"tour": "pga", "event_id": ev["event_id"], "year": ev["calendar_year"],
                    "file_format": "json", "key": api_key}, timeout=30,
        ).json()
        for p in detail.get("scores", []):
            rounds = [p[k] for k in ("round_1", "round_2", "round_3", "round_4")
                      if isinstance(p.get(k), dict)]
            rec = {}
            for stat in ("sg_ott", "sg_app", "sg_arg", "sg_putt", "sg_total"):
                vals = [r[stat] for r in rounds if r.get(stat) is not None]
                if vals:
                    rec[stat] = sum(vals) / len(vals)
            if rec:
                out[norm_name(p["player_name"])] = rec
        print(f"  archive {week}: {ev['event_name']} -> {len(detail.get('scores', []))} players")
    return out


def fetch_dg_points_by_week(api_key: str, weeks: list) -> dict:
    """dg_points per player for each week-ending Sunday, merging all events that week."""
    listing = requests.get(
        "https://feeds.datagolf.com/historical-event-data/event-list",
        params={"tour": "pga", "file_format": "json", "key": api_key}, timeout=30,
    ).json()
    by_week = {}
    for wk in weeks:
        by_week[wk] = {}
        for ev in listing:
            if ev.get("date") == wk:
                detail = requests.get(
                    "https://feeds.datagolf.com/historical-event-data/events",
                    params={"tour": "pga", "event_id": ev["event_id"], "year": ev["calendar_year"],
                            "file_format": "json", "key": api_key}, timeout=30,
                ).json()
                for p in detail.get("event_stats", []):
                    if p.get("dg_points") is not None:
                        by_week[wk][norm_name(p["player_name"])] = round(float(p["dg_points"]), 2)
                print(f"  dg_points {wk}: {ev['event_name']} -> {len(detail.get('event_stats', []))} players")
    return by_week


def fill_dg_points(ws, api_key: str) -> None:
    """Fill any completely-empty week columns in the DG Points sheet from the archive."""
    cols = date_columns(ws)
    player_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value not in (None, "")]
    empty = [(c, d) for c, d in cols
             if all(ws.cell(row=r, column=c).value in (None, "") for r in player_rows)]
    if not empty:
        print("[ok] DG Points: no empty week columns")
        return
    weeks = [d.date().isoformat() for _, d in empty]
    data = fetch_dg_points_by_week(api_key, weeks)
    filled = 0
    for c, d in empty:
        wk = d.date().isoformat()
        for r in player_rows:
            v = data.get(wk, {}).get(norm_name(ws.cell(row=r, column=1).value))
            if v is not None:
                ws.cell(row=r, column=c).value = v
                filled += 1
    print(f"[ok] DG Points: filled {filled} cells across {len(empty)} week(s)")


# PGA Database column -> player-decompositions field. These are the two far-right
# numbers on Data Golf's Course History tab (verified 149/149 and 147/149 against
# the workbook's hand-entered Scottish Open values).
COURSE_COLUMNS = {
    "T2Green": "total_fit_adjustment",
    "History": "total_course_history_adjustment",
}


def fill_course_fit(ws, api_key: str) -> None:
    """Write Data Golf's course fit + course history for the upcoming event into PGA Database."""
    d = requests.get(
        "https://feeds.datagolf.com/preds/player-decompositions",
        params={"tour": "pga", "file_format": "json", "key": api_key}, timeout=30,
    ).json()
    players = {norm_name(p["player_name"]): p for p in d.get("players", [])}
    print(f"course fit/history source: {d.get('event_name')} at {d.get('course_name')} ({len(players)} players)")
    for column_header, field in COURSE_COLUMNS.items():
        col = next((c.column for c in ws[1] if str(c.value).strip() == column_header), None)
        if col is None:
            print(f"[warn] column {column_header!r} not found in {ws.title}, skipping")
            continue
        hits = 0
        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            if name in (None, ""):
                continue
            rec = players.get(norm_name(name))
            v = rec.get(field) if rec else None
            if v is not None:
                ws.cell(row=r, column=col).value = round(float(v), 2)
                hits += 1
            else:
                ws.cell(row=r, column=col).value = None  # not in this week's field: clear stale value
        print(f"[ok] {ws.title}.{column_header} <- {field}: set for {hits} field players, others cleared")


FIELD_HEADERS = [
    "Player", "SG OTT", "Points", "Approach", "Putting", "Around Green",
    "T2Green", "Form", "History", "Tournament", "Tour",
]


def _header_fill():
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor="00B050")


def _header_font():
    from openpyxl.styles import Font
    return Font(name="Aptos Narrow", size=11, bold=True)


def _body_font():
    from openpyxl.styles import Font
    return Font(name="Aptos Narrow", size=11)


def build_event_field_sheet(
    wb,
    sheet_name: str,
    tournament: str,
    tour: str = "PGA Tour",
) -> int:
    """Build an event field sheet the same way Scottish Open Field is structured.

    Takes every PGA Database row with a non-null T2Green (this week's field),
    copies the rolling SG/Form formulas (pointing at the same source rows),
    writes T2Green/History values, and sets Tournament/Tour. Replaces any
    existing sheet of the same name. Also updates PGA Database.Tournament for
    those field players so the master sheet stays in sync.
    """
    if "PGA Database" not in wb.sheetnames:
        raise SystemExit("PGA Database sheet required to build a field sheet")
    db = wb["PGA Database"]
    points_sheet = "Class" if "Class" in wb.sheetnames else (
        "DG Points" if "DG Points" in wb.sheetnames else None
    )
    if points_sheet is None:
        raise SystemExit("need Class or DG Points sheet for Points formulas")

    # Source-row formulas: same pattern as Scottish Open Field / PGA Database.
    def formula_for(col: int, src_row: int) -> str:
        if col == 2:
            return f"='SG OTT'!Q{src_row}"
        if col == 3:
            ref = f"'{points_sheet}'!Q{src_row}" if " " in points_sheet else f"{points_sheet}!Q{src_row}"
            return f"={ref}"
        if col == 4:
            return f"='SG APP'!Q{src_row}"
        if col == 5:
            return f"='SG PUTT'!Q{src_row}"
        if col == 6:
            return f"='SG ARG'!Q{src_row}"
        if col == 8:
            return f"='Current Form'!L{src_row}"
        raise ValueError(col)

    field_rows = []
    for r in range(2, db.max_row + 1):
        name = db.cell(row=r, column=1).value
        t2g = db.cell(row=r, column=7).value
        if name in (None, "") or t2g in (None, ""):
            continue
        field_rows.append(r)
        db.cell(row=r, column=10).value = tournament  # Tournament column

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    # Place next to Scottish Open Field when present; else after PGA Database.
    anchor = "Scottish Open Field" if "Scottish Open Field" in wb.sheetnames else "PGA Database"
    ws = wb.create_sheet(sheet_name, wb.sheetnames.index(anchor) + 1)

    header_fill = _header_fill()
    header_font = _header_font()
    body_font = _body_font()
    number_formats = {
        2: "0.0000", 3: "0.0000", 4: "0.0000", 5: "0.0000", 6: "0.0000",
        7: "0.00", 8: "0.00", 9: "0.00",
    }
    col_widths = {
        "A": 27.0, "B": 9.71, "C": 9.0, "D": 9.43, "E": 9.14, "F": 13.57,
        "G": 9.14, "H": 9.14, "I": 9.14, "J": 28.72, "K": 9.14,
    }

    for c, header in enumerate(FIELD_HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill = header_fill
        cell.font = header_font
        if c in number_formats:
            cell.number_format = number_formats[c]
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"

    for out_i, src_row in enumerate(field_rows, start=2):
        ws.cell(row=out_i, column=1, value=db.cell(row=src_row, column=1).value).font = body_font
        for col in (2, 3, 4, 5, 6, 8):
            # Same formula pattern as Scottish Open Field / PGA Database.
            cell = ws.cell(row=out_i, column=col, value=formula_for(col, src_row))
            cell.font = body_font
            cell.number_format = number_formats[col]
        for col in (7, 9):  # T2Green, History — literal values from PGA Database
            cell = ws.cell(row=out_i, column=col, value=db.cell(row=src_row, column=col).value)
            cell.font = body_font
            cell.number_format = number_formats[col]
        ws.cell(row=out_i, column=10, value=tournament).font = body_font
        tour_val = db.cell(row=src_row, column=11).value or tour
        ws.cell(row=out_i, column=11, value=tour_val).font = body_font

    print(f"[ok] {sheet_name}: {len(field_rows)} field players "
          f"(Tournament={tournament!r}, formulas -> {points_sheet}/SG*/Current Form)")
    return len(field_rows)


def export_field_csv_from_sheet(wb, sheet_name: str, csv_path: str) -> int:
    """Export a field-stats CSV from an event field sheet (values already resolved)."""
    import csv

    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"sheet {sheet_name!r} not found for CSV export")
    ws = wb[sheet_name]
    rows_out = [FIELD_HEADERS]
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name in (None, ""):
            continue
        rows_out.append([ws.cell(row=r, column=c).value for c in range(1, 12)])
    os.makedirs(os.path.dirname(csv_path) or ".", exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows_out)
    print(f"[ok] wrote {csv_path} ({len(rows_out) - 1} players)")
    return len(rows_out) - 1


def export_field_csv_from_database(workbook_path: str, csv_path: str, tournament: str) -> int:
    """Export field-stats CSV from cached PGA Database values (T2Green field filter)."""
    import csv

    db = openpyxl.load_workbook(workbook_path, data_only=True)["PGA Database"]
    rows_out = [FIELD_HEADERS]
    for r in range(2, db.max_row + 1):
        if db.cell(row=r, column=7).value in (None, ""):
            continue
        row = [db.cell(row=r, column=c).value for c in range(1, 12)]
        row[9] = tournament  # Tournament
        if not row[10]:
            row[10] = "PGA Tour"
        # Skip rows whose SG aggregates were stripped (openpyxl save without recalc)
        if all(row[c] in (None, "") for c in range(1, 6)):
            continue
        rows_out.append(row)
    os.makedirs(os.path.dirname(csv_path) or ".", exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows_out)
    print(f"[ok] wrote {csv_path} ({len(rows_out) - 1} players)")
    return len(rows_out) - 1


def date_columns(ws) -> list:
    """(column_index, datetime) pairs for the date headers in row 1."""
    cols = []
    for cell in ws[1]:
        if isinstance(cell.value, datetime):
            cols.append((cell.column, cell.value))
    return cols


def update_sheet(ws, new_weeks: list, stat: str) -> tuple:
    """Shift the sheet's date window and write new week values. Returns (hits, misses)."""
    cols = date_columns(ws)
    if not cols:
        raise SystemExit(f"{ws.title}: no date headers found in row 1")
    col_indexes = [c for c, _ in cols]
    existing_dates = [d for _, d in cols]

    all_dates = existing_dates + [d for d, _ in new_weeks if d not in existing_dates]
    all_dates.sort()
    window = all_dates[-len(col_indexes):]

    player_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value not in (None, "")]

    # Snapshot current values by (player, date) before overwriting.
    current = {}
    for r in player_rows:
        for (c, d) in cols:
            current[(r, d)] = ws.cell(row=r, column=c).value
    new_data = {d: values for d, values in new_weeks}

    hits = misses = 0
    matched_names = set()
    for r in player_rows:
        key = norm_name(ws.cell(row=r, column=1).value)
        for c, d in zip(col_indexes, window):
            ws.cell(row=1, column=c).value = d
            if d in new_data:
                if stat is None:
                    value = None
                else:
                    rec = new_data[d].get(key)
                    value = None if rec is None or rec.get(stat) is None else round(float(rec[stat]), 2)
                    if rec is not None:
                        matched_names.add(key)
            else:
                value = current.get((r, d))
            ws.cell(row=r, column=c).value = value
    if stat is not None and new_weeks:
        all_players = set().union(*(set(v) for _, v in new_weeks))
        hits = len(matched_names)
        misses = len(all_players - matched_names)
    return hits, misses


def main() -> int:
    parser = argparse.ArgumentParser(description="Update StatCaddy workbook from Data Golf.")
    parser.add_argument("workbook")
    parser.add_argument("output")
    parser.add_argument("--snapshot", help="pull_all.py snapshot dir for a completed event")
    parser.add_argument("--snapshot-date", help="Week (Sunday) YYYY-MM-DD for the snapshot event")
    parser.add_argument("--live", action="store_true", help="Fetch current event live from Data Golf")
    parser.add_argument("--live-date", help="Week (Sunday) YYYY-MM-DD for the live event")
    parser.add_argument("--archive-date", action="append", default=[],
                        help="Week (Sunday) YYYY-MM-DD to fill/refresh from the official archive (repeatable)")
    parser.add_argument("--dg-points", action="store_true",
                        help="Fill empty DG Points week columns from the historical archive (needs historical tier)")
    parser.add_argument("--course-fit", action="store_true",
                        help="Refresh PGA Database T2Green with course fit for the upcoming event")
    parser.add_argument("--field-sheet",
                        help="Build/replace an event field sheet (e.g. 'Wyndham Field') "
                             "from PGA Database rows with T2Green, matching Scottish Open Field")
    parser.add_argument("--tournament",
                        help="Tournament label for --field-sheet / --export-field-csv "
                             "(default: sheet name without trailing ' Field')")
    parser.add_argument("--export-field-csv",
                        help="Write a Scottish-Open-Field-style CSV of the current T2Green field")
    args = parser.parse_args()

    new_weeks = []
    if args.snapshot:
        if not args.snapshot_date:
            parser.error("--snapshot requires --snapshot-date")
        new_weeks.append((datetime.fromisoformat(args.snapshot_date), week_from_snapshot(args.snapshot)))
    if args.live:
        if not args.live_date:
            parser.error("--live requires --live-date")
        if not os.environ.get("DATAGOLF_KEY"):
            print("ERROR: DATAGOLF_KEY not set", file=sys.stderr)
            return 2
        new_weeks.append((datetime.fromisoformat(args.live_date), week_from_live(os.environ["DATAGOLF_KEY"])))
    for wk in args.archive_date:
        data = week_from_archive(os.environ["DATAGOLF_KEY"], wk)
        if not data:
            print(f"ERROR: archive has no events completing {wk} yet — aborting so the "
                  f"workbook is not written with an empty week.", file=sys.stderr)
            return 3
        new_weeks.append((datetime.fromisoformat(wk), data))
    if not new_weeks and not (args.dg_points or args.course_fit or args.field_sheet or args.export_field_csv):
        parser.error("nothing to do: pass --snapshot/--live and/or --dg-points/--course-fit/"
                     "--field-sheet/--export-field-csv")
    empty = [d for d, v in new_weeks if not v]
    for d in empty:
        print(f"[skip] {d.date()}: no event data for that week (off week?) — not adding a column")
    new_weeks = [(d, v) for d, v in new_weeks if v]
    new_weeks.sort(key=lambda x: x[0])

    wb = openpyxl.load_workbook(args.workbook, data_only=False)
    if new_weeks:
        for sheet_name, stat in STAT_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                print(f"[warn] sheet {sheet_name!r} not found, skipping")
                continue
            hits, misses = update_sheet(wb[sheet_name], new_weeks, stat)
            note = "(window shifted)" if stat is None else f"matched {hits} sheet players; {misses} DG players not in sheet"
            print(f"[ok] {sheet_name}: {note}")

    if args.dg_points:
        fill_dg_points(wb["Class"], os.environ["DATAGOLF_KEY"])
    if args.course_fit:
        fill_course_fit(wb["PGA Database"], os.environ["DATAGOLF_KEY"])

    tournament = args.tournament
    if args.field_sheet:
        if not tournament:
            tournament = args.field_sheet.removesuffix(" Field").removesuffix(" field").strip()
            if not tournament:
                parser.error("--field-sheet requires --tournament when the sheet name has no ' Field' suffix")
        build_event_field_sheet(wb, args.field_sheet, tournament)

    wb.save(args.output)
    print(f"\nSaved: {args.output}")

    # openpyxl drops cached formula results on save; LibreOffice restores them so
    # data_only readers / Excel Online previews show SG aggregates again.
    if _recalc_with_libreoffice(args.output):
        print(f"[ok] recalculated formula caches via LibreOffice: {args.output}")

    if args.export_field_csv:
        if args.field_sheet:
            # Reload post-recalc so CSV gets resolved formula values.
            export_field_csv_from_sheet(
                openpyxl.load_workbook(args.output, data_only=True),
                args.field_sheet,
                args.export_field_csv,
            )
        else:
            if not tournament:
                parser.error("--export-field-csv requires --tournament (or --field-sheet)")
            export_field_csv_from_database(args.output, args.export_field_csv, tournament)
    return 0


def _recalc_with_libreoffice(path: str) -> bool:
    """Headless LibreOffice recalc to restore cached formula values. Returns True on success."""
    import shutil
    import subprocess
    import tempfile

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        print("[warn] LibreOffice not found — formula caches not recalculated "
              "(open in Excel once to populate values)")
        return False
    path = os.path.abspath(path)
    outdir = tempfile.mkdtemp(prefix="xlsx-recalc-")
    try:
        # Convert to xlsx in a temp dir (forces full recalc), then replace original.
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", outdir, path],
            check=True, capture_output=True, timeout=300,
        )
        converted = os.path.join(outdir, os.path.basename(path))
        if not os.path.isfile(converted):
            # LibreOffice sometimes alters the extension casing / name
            candidates = [f for f in os.listdir(outdir) if f.lower().endswith(".xlsx")]
            if not candidates:
                print(f"[warn] LibreOffice produced no xlsx in {outdir}")
                return False
            converted = os.path.join(outdir, candidates[0])
        shutil.move(converted, path)
        return True
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
        print(f"[warn] LibreOffice recalc failed: {exc}")
        return False
    finally:
        shutil.rmtree(outdir, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
